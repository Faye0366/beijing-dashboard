# -*- coding: utf-8 -*-
"""
宽带业务数据看板 - 数据处理脚本
读取 Excel "流量套餐数据源" sheet，筛选 E列=宽带，聚合成功订单数据，生成自包含 HTML。
成功订单(J列)已包含"成功"与"处理中"两种状态，直接取值即可。

用法: python process_data_broadband.py
输出: broadband.html (看板), broadband_compare.html (对比页)
"""

import json
import os
from datetime import datetime
from collections import defaultdict
import openpyxl

# ===== 配置 =====
EXCEL_PATH = r"C:\Users\Faye\Desktop\WorkBuddy数据看板\北京业务数据统计.xlsx"
OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))
SHEET_NAME = "流量套餐数据源"
HTML_TEMPLATE = os.path.join(OUTPUT_DIR, "templates", "broadband.html")
FINAL_HTML = os.path.join(OUTPUT_DIR, "broadband.html")
COMPARE_TEMPLATE = os.path.join(OUTPUT_DIR, "templates", "broadband_compare.html")
COMPARE_HTML = os.path.join(OUTPUT_DIR, "broadband_compare.html")

BIZ_TYPES = ["宽带新装", "宽带续费", "宽带提速"]


def main():
    print(f"正在读取: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb[SHEET_NAME]

    # ===== 聚合容器 =====
    monthly = defaultdict(lambda: {"total": 0, "公域": 0, "私域": 0})
    daily = defaultdict(lambda: defaultdict(lambda: {"total": 0, "公域": 0, "私域": 0}))
    biz_monthly = defaultdict(lambda: {b: 0 for b in BIZ_TYPES})
    public_top5 = defaultdict(lambda: defaultdict(int))   # month -> product(B) -> count
    private_top5 = defaultdict(lambda: defaultdict(int))
    biz_top3 = defaultdict(lambda: {b: defaultdict(int) for b in BIZ_TYPES})  # month -> biz -> product(B) -> count
    # 对比明细: product_detail[date][B产品名][U域][V点位] = value
    product_detail = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(int))))
    # 业务类型对比明细: biz_detail[date][T业务类型][U域] = value
    biz_detail = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
    all_products = set()
    all_dates = set()

    row_count = 0
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        e_val = row[4].value                              # E: 类型
        if e_val != "宽带":
            continue
        date_val = row[0].value                           # A: 日期
        if date_val is None or not hasattr(date_val, "year"):
            continue
        success_val = row[9].value                        # J: 成功订单
        b_val = row[1].value                             # B: 产品名称
        t_val = row[19].value                            # T: 统一产品名称（业务类型）
        u_val = row[20].value                            # U: 公/私域
        v_val = row[21].value                            # V: 投放点位

        # 处理成功订单值（J列已含"成功"与"处理中"）
        if success_val is None:
            success = 0
        elif isinstance(success_val, (int, float)):
            success = int(success_val)
        else:
            try:
                success = int(success_val)
            except (ValueError, TypeError):
                success = 0

        # 公/私域分类（U列为空的25行不计入公域/私域，但计入总单量）
        if u_val is None or (isinstance(u_val, str) and u_val.strip() == ""):
            u = None
        else:
            u = u_val if u_val in ("公域", "私域") else "外部渠道"

        v = v_val if v_val else "未知"
        t = t_val if t_val in BIZ_TYPES else "其他"
        month_key = f"{date_val.year}-{date_val.month:02d}"
        day_key = date_val.day
        date_str = date_val.strftime("%Y-%m-%d")
        product = str(b_val) if b_val else "未知"

        # 月度 / 日度汇总
        monthly[month_key]["total"] += success
        daily[month_key][day_key]["total"] += success
        if u in ("公域", "私域"):
            monthly[month_key][u] += success
            daily[month_key][day_key][u] += success

        # 业务类型月度
        if t in BIZ_TYPES:
            biz_monthly[month_key][t] += success

        # TOP5 产品（按 B 列，按月，按 U 域）
        if u == "公域":
            public_top5[month_key][product] += success
        elif u == "私域":
            private_top5[month_key][product] += success

        # 业务类型 TOP3 产品（按 B 列，按月）
        if t in BIZ_TYPES:
            biz_top3[month_key][t][product] += success

        # 对比明细
        if u in ("公域", "私域"):
            product_detail[date_str][product][u][v] += success
            biz_detail[date_str][t][u] += success
            all_products.add(product)
        all_dates.add(date_str)

        row_count += 1

    print(f"处理完成，共 {row_count} 行数据")

    # ===== 整理输出格式 =====
    months = sorted(monthly.keys())

    monthly_data = [{
        "month": m,
        "total": monthly[m]["total"],
        "公域": monthly[m]["公域"],
        "私域": monthly[m]["私域"],
    } for m in months]

    daily_data = {}
    for m in months:
        days = sorted(daily[m].keys())
        daily_data[m] = {
            "days": [d for d in days],
            "total": [daily[m][d]["total"] for d in days],
            "公域": [daily[m][d]["公域"] for d in days],
            "私域": [daily[m][d]["私域"] for d in days],
        }

    biz_monthly_data = {m: dict(biz_monthly[m]) for m in months}

    def topn(src, n):
        return [{"name": k, "value": v} for k, v in sorted(src.items(), key=lambda x: x[1], reverse=True)[:n]]

    public_top5_data = {m: topn(public_top5[m], 5) for m in months}
    private_top5_data = {m: topn(private_top5[m], 5) for m in months}
    biz_top3_data = {m: {b: topn(biz_top3[m][b], 3) for b in BIZ_TYPES} for m in months}

    result = {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "current_month": months[-1] if months else "",
        "months": months,
        "biz_types": BIZ_TYPES,
        "monthly_data": monthly_data,
        "daily_data": daily_data,
        "biz_monthly": biz_monthly_data,
        "public_top5": public_top5_data,
        "private_top5": private_top5_data,
        "biz_top3": biz_top3_data,
    }

    # ===== 生成看板 HTML =====
    json_str = json.dumps(result, ensure_ascii=False)
    with open(HTML_TEMPLATE, "r", encoding="utf-8") as f:
        html = f.read()
    html = html.replace("// ${DATA_PLACEHOLDER}", "const DASHBOARD_DATA = " + json_str + ";")
    with open(FINAL_HTML, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"看板已生成: {FINAL_HTML} ({os.path.getsize(FINAL_HTML) / 1024:.1f} KB)")

    # ===== 生成对比页 HTML =====
    compare_result = {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "all_products": sorted(all_products),
        "all_dates": sorted(all_dates),
        "months": months,
        "biz_types": BIZ_TYPES,
        "product_detail": {
            d: {p: {u: dict(ch) for u, ch in pd.items()} for p, pd in product_detail[d].items()}
            for d in sorted(product_detail)
        },
        "biz_detail": {
            d: {t: dict(u) for t, u in biz_detail[d].items()}
            for d in sorted(biz_detail)
        },
    }
    compare_json = json.dumps(compare_result, ensure_ascii=False)
    with open(COMPARE_TEMPLATE, "r", encoding="utf-8") as f:
        chtml = f.read()
    chtml = chtml.replace("// ${DATA_PLACEHOLDER}", "const COMPARE_DATA = " + compare_json + ";")
    with open(COMPARE_HTML, "w", encoding="utf-8") as f:
        f.write(chtml)
    print(f"对比页已生成: {COMPARE_HTML} ({os.path.getsize(COMPARE_HTML) / 1024:.1f} KB)")

    # ===== 摘要 =====
    print("\n=== 宽带月度摘要 ===")
    for m in monthly_data:
        print(f"  {m}: 总={m['total']}, 公域={m['公域']}, 私域={m['私域']}")
    print("\n=== 业务类型(最新月) ===")
    if months:
        lm = months[-1]
        print(f"  {lm}: {biz_monthly_data[lm]}")


if __name__ == "__main__":
    main()
