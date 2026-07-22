# -*- coding: utf-8 -*-
"""
流量套餐数据看板 - 数据处理脚本
读取 Excel "流量套餐数据源" sheet，聚合成功订单数据，生成自包含的 HTML 看板文件。

用法: python process_data.py
输出: 流量套餐订单看板.html (直接用浏览器打开即可查看)
"""

import json
import os
from datetime import datetime
from collections import defaultdict
import openpyxl

# ===== 配置 =====
# Excel 文件路径（桌面）
EXCEL_PATH = r"C:\Users\Faye\Desktop\WorkBuddy数据看板\北京业务数据统计.xlsx"
OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))
SHEET_NAME = "流量套餐数据源"
HTML_TEMPLATE = os.path.join(OUTPUT_DIR, "templates", "index.html")
FINAL_HTML = os.path.join(OUTPUT_DIR, "index.html")
COMPARE_TEMPLATE = os.path.join(OUTPUT_DIR, "templates", "compare.html")
COMPARE_HTML = os.path.join(OUTPUT_DIR, "compare.html")


def main():
    print(f"正在读取: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb[SHEET_NAME]

    # ===== 读取数据 =====
    # 数据结构:
    # A=日期, B=产品名称, ..., J=成功订单, ..., U=公/私域, V=投放点位

    monthly_orders = defaultdict(lambda: {"total": 0, "公域": 0, "私域": 0, "外部渠道": 0})
    daily_orders = defaultdict(lambda: defaultdict(lambda: {"total": 0, "公域": 0, "私域": 0, "外部渠道": 0}))
    # monthly_channel_orders[month][投放点位][公/私域] = 成功订单量
    monthly_channel_orders = defaultdict(lambda: defaultdict(lambda: {"公域": 0, "私域": 0, "外部渠道": 0, "total": 0}))
    # channel_product_orders[month][投放点位][产品名称] = 成功订单量
    channel_product_orders = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
    # 充流量频道产品月度数据 [month][product_name] = 成功订单量
    cll_product_monthly = defaultdict(lambda: defaultdict(int))
    # 充流量频道产品日期级数据 [date_str][product_name] = 成功订单量（用于对比页面）
    cll_product_daily = defaultdict(lambda: defaultdict(int))
    # 公域渠道点位日期级数据 [date_str][投放点位][具体投放位置] = 成功订单量（用于对比页面）
    channel_public_daily = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
    # 私域渠道点位日期级数据 [date_str][投放点位] = 成功订单量（用于对比页面）
    channel_private_daily = defaultdict(lambda: defaultdict(int))
    # product_detail[date_str][统一产品名称(T列)][公/私域(U列)][投放点位(V列)] = 成功订单量
    product_detail = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(int))))
    all_products_set = set()
    all_dates_set = set()

    row_count = 0
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        date_val = row[0].value    # A: 日期
        success_val = row[9].value  # J: 成功订单
        e_val = row[4].value        # E: 类型
        b_val = row[1].value        # B: 产品名称
        t_val = row[19].value       # T: 统一产品名称
        u_val = row[20].value       # U: 公/私域
        v_val = row[21].value       # V: 投放点位
        w_val = row[22].value if len(row) > 22 else None  # W: 具体投放位置

        if date_val is None or not hasattr(date_val, "year"):
            continue

        # E列"类型"为"宽带"或"号卡"的不计入流量套餐统计
        if e_val in ("宽带", "号卡"):
            continue

        # 处理成功订单值
        if success_val is None:
            success = 0
        elif isinstance(success_val, (int, float)):
            success = int(success_val)
        else:
            try:
                success = int(success_val)
            except (ValueError, TypeError):
                success = 0

        # U列(公/私域)为空的不计入任何统计
        if u_val is None or (isinstance(u_val, str) and u_val.strip() == ""):
            continue

        month_key = f"{date_val.year}-{date_val.month:02d}"
        day_key = date_val.day
        date_str = date_val.strftime("%Y-%m-%d")

        # 公/私域分类，默认归为"外部渠道"以外的处理
        domain = u_val if u_val in ("公域", "私域", "外部渠道") else "外部渠道"

        # 月度汇总
        monthly_orders[month_key]["total"] += success
        monthly_orders[month_key][domain] += success

        # 日度汇总
        daily_orders[month_key][day_key]["total"] += success
        daily_orders[month_key][day_key][domain] += success

        # 投放点位汇总（合并"承接页-话费专区"、"承接页-流量专区"、"承接页-语音专区"为"承接页"）
        channel = v_val if v_val else "未知"
        if channel in ("承接页-话费专区", "承接页-流量专区", "承接页-语音专区"):
            channel = "承接页"
        monthly_channel_orders[month_key][channel]["total"] += success
        monthly_channel_orders[month_key][channel][domain] += success

        # 投放点位×产品 汇总（用于 TOP5 产品图表）
        product_name = b_val if b_val else "未知"
        channel_product_orders[month_key][channel][product_name] += success

        # 充流量频道产品月度数据（用于充流量TOP10模块）
        if channel == "充流量":
            cll_product_monthly[month_key][product_name] += success
            cll_product_daily[date_str][product_name] += success

        # 渠道点位对比数据（仅流量包和套餐，用于对比页面）
        if e_val in ("流量包", "套餐"):
            if domain == "公域":
                location = w_val if w_val else "未指定"
                channel_public_daily[date_str][channel][location] += success
            elif domain == "私域":
                channel_private_daily[date_str][channel] += success

        # 产品明细（用于数据对比页面）— 使用 T列统一产品名称
        unified_name = t_val if t_val else (b_val if b_val else "未知")
        # 过滤掉非实际产品的条目（页面/UV等）
        EXCLUDE_KEYWORDS = ("小程序首页", "访问页", "通用页", "查询页", "UV", "访问量")
        if not any(kw in unified_name for kw in EXCLUDE_KEYWORDS):
            product_detail[date_str][unified_name][domain][channel] += success
            all_products_set.add(unified_name)
        all_dates_set.add(date_str)

        row_count += 1

    print(f"处理完成，共 {row_count} 行数据")

    # ===== 整理输出格式 =====
    months = sorted(monthly_orders.keys())

    # 1. 月度订单量
    monthly_data = []
    for m in months:
        monthly_data.append({
            "month": m,
            "total": monthly_orders[m]["total"],
            "公域": monthly_orders[m]["公域"],
            "私域": monthly_orders[m]["私域"],
            "外部渠道": monthly_orders[m]["外部渠道"],
        })

    # 2. 日度订单量 - 按月份组织
    daily_data = {}
    for m in months:
        days = sorted(daily_orders[m].keys())
        daily_data[m] = {
            "days": [d for d in days],
            "total": [daily_orders[m][d]["total"] for d in days],
            "公域": [daily_orders[m][d]["公域"] for d in days],
            "私域": [daily_orders[m][d]["私域"] for d in days],
            "外部渠道": [daily_orders[m][d]["外部渠道"] for d in days],
        }

    # 3. 投放点位月度订单量
    all_channels = set()
    for m in months:
        all_channels.update(monthly_channel_orders[m].keys())
    all_channels = sorted(all_channels)

    channel_data = {}
    for m in months:
        channel_data[m] = []
        for ch in all_channels:
            d = monthly_channel_orders[m].get(ch, {"公域": 0, "私域": 0, "外部渠道": 0, "total": 0})
            channel_data[m].append({
                "channel": ch,
                "公域": d["公域"],
                "私域": d["私域"],
                "外部渠道": d["外部渠道"],
                "total": d["total"],
            })

    # 4. 各投放点位 TOP5 产品
    channel_top_products = {}
    for m in months:
        channel_top_products[m] = {}
        for ch in all_channels:
            products = channel_product_orders[m].get(ch, {})
            # 按订单量降序取 TOP5
            top5 = sorted(products.items(), key=lambda x: x[1], reverse=True)[:5]
            channel_top_products[m][ch] = [{"name": name, "value": val} for name, val in top5]

    # 当前月份（数据中最新月份）
    current_month = months[-1] if months else ""

    result = {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "current_month": current_month,
        "months": months,
        "channels": all_channels,
        "monthly_data": monthly_data,
        "daily_data": daily_data,
        "channel_data": channel_data,
        "channel_top_products": channel_top_products,
        "cll_product_monthly": {m: dict(v) for m, v in sorted(cll_product_monthly.items())},
    }

    # ===== 生成自包含 HTML 看板 =====
    json_str = json.dumps(result, ensure_ascii=False)

    with open(HTML_TEMPLATE, "r", encoding="utf-8") as f:
        html = f.read()

    html = html.replace("// ${DATA_PLACEHOLDER}", "const DASHBOARD_DATA = " + json_str + ";")

    with open(FINAL_HTML, "w", encoding="utf-8") as f:
        f.write(html)

    print(f"\n看板已生成: {FINAL_HTML}")
    print(f"文件大小: {os.path.getsize(FINAL_HTML) / 1024:.1f} KB")
    print(f"直接用浏览器打开即可查看，或推送到 GitHub Pages。")

    # 打印摘要
    print("\n=== 月度订单量摘要 ===")
    for m in monthly_data:
        print(f"  {m['month']}: 总计={m['total']}, 公域={m['公域']}, 私域={m['私域']}, 外部={m['外部渠道']}")

    # ===== 生成数据对比页面 =====
    compare_result = {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "all_products": sorted(all_products_set),
        "all_dates": sorted(all_dates_set),
        "all_channels": all_channels,
        "months": months,
        "product_detail": {d: dict(v) for d, v in sorted(product_detail.items())},
        "cll_product_daily": {d: dict(v) for d, v in sorted(cll_product_daily.items())},
        "channel_public_daily": {d: {ch: dict(locs) for ch, locs in v.items()} for d, v in sorted(channel_public_daily.items())},
        "channel_private_daily": {d: dict(v) for d, v in sorted(channel_private_daily.items())},
    }

    compare_json = json.dumps(compare_result, ensure_ascii=False)

    with open(COMPARE_TEMPLATE, "r", encoding="utf-8") as f:
        compare_html = f.read()

    compare_html = compare_html.replace("// ${DATA_PLACEHOLDER}", "const COMPARE_DATA = " + compare_json + ";")

    with open(COMPARE_HTML, "w", encoding="utf-8") as f:
        f.write(compare_html)

    print(f"\n对比页面已生成: {COMPARE_HTML}")
    print(f"文件大小: {os.path.getsize(COMPARE_HTML) / 1024:.1f} KB")
    print(f"直接用浏览器打开即可查看，或推送到 GitHub Pages。")


if __name__ == "__main__":
    main()
